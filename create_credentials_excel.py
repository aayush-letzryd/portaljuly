import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define path to Downloads folder
user_home = os.path.expanduser("~")
downloads_folder = os.path.join(user_home, "Downloads")
output_file = os.path.join(downloads_folder, "LetzRyd_Portal_Primary_Credentials.xlsx")

# Create workbook and select active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Primary Credentials"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Title Block
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = "LetzRyd Web Portal  —  Primary Login Credentials & User IDs"
title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
title_cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 35

# Sub-header Block
ws.merge_cells("A2:F2")
sub_cell = ws["A2"]
sub_cell.value = "Complete list of ALL 15 primary portal user accounts. Default Password for ALL accounts is: 123456"
sub_cell.font = Font(name="Calibri", size=10, italic=True, color="475569")
sub_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 22

# Table Headers
headers = [
    "User ID",
    "Username",
    "Role Code",
    "Role Name",
    "Generic Employee Title",
    "Default Password"
]

header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.row_dimensions[4].height = 28
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

# Complete list of all 15 portal accounts
data = [
    [3, "admin", "SA", "Super Admin", "System Super Admin", "123456"],
    [41, "general_manager", "GM", "General Manager", "General Manager 1", "123456"],
    [17, "business", "BH", "Business Head", "Business Head 1", "123456"],
    [18, "finance_lead", "FL", "Finance Lead", "Finance Lead 1", "123456"],
    [19, "finance_executive", "FE", "Finance Executive", "Finance Executive 1", "123456"],
    [20, "city_manager", "CM", "City Manager", "City Manager 1", "123456"],
    [21, "driver_manager", "DM", "Driver Manager", "Driver Manager 1", "123456"],
    [23, "ops_executive", "OE", "Ops Executive", "Ops Executive 1", "123456"],
    [24, "fleet_manager", "FM", "Fleet Manager", "Fleet Manager 1", "123456"],
    [25, "maintenance_coordinator", "MC", "Maintenance Coordinator", "Maintenance Coordinator 1", "123456"],
    [26, "onboarding_executive", "OB", "Onboarding Executive", "Onboarding Executive 1", "123456"],
    [28, "support_executive", "SP", "Support Executive", "Support Executive 1", "123456"],
    [29, "auditor", "AU", "Auditor / Compliance", "Auditor 1", "123456"],
    [31, "fleet_partner", "PT", "Fleet Partner / Operator", "Fleet Partner 1", "123456"],
    [32, "driver", "DR", "Driver", "Driver 1", "123456"]
]

# Styling definitions for data cells
font_regular = Font(name="Calibri", size=10, color="0F172A")
font_bold_center = Font(name="Calibri", size=10, bold=True, color="0F172A")
font_user = Font(name="Calibri", size=10, bold=True, color="047857")
font_pwd = Font(name="Calibri", size=10, bold=True, color="B45309")

fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
fill_odd = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0")
)

start_row = 5
for row_idx, row_data in enumerate(data, start=start_row):
    row_fill = fill_odd if row_idx % 2 == 1 else fill_even
    ws.row_dimensions[row_idx].height = 26
    
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = val
        cell.fill = row_fill
        cell.border = thin_border
        
        if col_idx in [1, 3, 6]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
        if col_idx == 2: # Username
            cell.font = font_user
        elif col_idx == 6: # Password
            cell.font = font_pwd
        elif col_idx in [1, 3]: # User ID, Role Code
            cell.font = font_bold_center
        else:
            cell.font = font_regular

# Column Widths
col_widths = {
    "A": 12, # User ID
    "B": 26, # Username
    "C": 14, # Role Code
    "D": 26, # Role Name
    "E": 28, # Generic Employee Title
    "F": 20  # Default Password
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save file to Downloads folder
wb.save(output_file)
print(f"[OK] Complete 15-Role Credentials Excel file saved to Downloads: {output_file}")
