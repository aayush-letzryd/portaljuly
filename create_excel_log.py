import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define path to Downloads folder
user_home = os.path.expanduser("~")
downloads_folder = os.path.join(user_home, "Downloads")
output_file = os.path.join(downloads_folder, "LetzRyd_Portal_Form_Feedback_and_Implementation_Log.xlsx")

# Create workbook and select active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Feature Implementation Log"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Title Block
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "LetzRyd Web Portal  —  Complete Functional Features & Field Implementation Log"
title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
title_cell.fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 35

# Sub-header Block
ws.merge_cells("A2:H2")
sub_cell = ws["A2"]
sub_cell.value = "Comprehensive list of input fields, document uploads, validation rules, REST endpoints & workflow capabilities added across all forms"
sub_cell.font = Font(name="Calibri", size=10, italic=True, color="475569")
sub_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 22

# Table Headers
headers = [
    "Date",
    "Item #",
    "Feature / Requirement Added",
    "Requested By",
    "Implemented By",
    "Status",
    "Location",
    "Functional Change & Implementation Details (Simple Language)"
]

header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.row_dimensions[4].height = 28
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=col_idx)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

# Full functional data list
data = [
    # Onboarding & KYC
    [
        "2026-08-04", 1,
        "Driver Selfie Photo Upload",
        "Vivek", "Anurag", "Done", "Onboarding Form",
        "Added mandatory driver selfie photo upload field using live camera capture or image selection during onboarding."
    ],
    [
        "2026-08-04", 2,
        "Aadhaar Card Front & Back Photo Uploads",
        "Vivek", "Anurag", "Done", "Onboarding Form",
        "Added separate upload fields for Aadhaar Card Front Photo and Aadhaar Card Back Photo with image preview."
    ],
    [
        "2026-08-04", 3,
        "PAN Card Upload & Aadhaar Link Checkbox",
        "Vivek", "Anurag", "Done", "Onboarding Form",
        "Added PAN Card Photo Upload field and a checkbox toggle to verify 'Is Aadhaar linked with PAN'."
    ],
    [
        "2026-08-05", 4,
        "Multi-Page Local Address Proof Upload (1 to 4 pages)",
        "Vivek", "Anurag", "Done", "Onboarding Form",
        "Added multi-page upload capability (pages 1-4) for local address proof documents such as rent agreement or utility bills."
    ],
    [
        "2026-08-04", 5,
        "Permanent & Local Address Input Fields",
        "Vivek", "Anurag", "Done", "Onboarding Form",
        "Added full address fields including House No, Street, Landmark, City, State, and Pincode for both local and permanent addresses."
    ],
    [
        "2026-08-04", 6,
        "Bank Account Details & Passbook Photo Upload",
        "Vivek", "Anurag", "Done", "Onboarding Form",
        "Added Bank Account Number, IFSC Code, Account Holder Name, and Passbook/Cancelled Cheque photo upload."
    ],
    [
        "2026-08-04", 7,
        "Emergency Contact Name, Phone & Relationship",
        "Vivek", "Anurag", "Done", "Onboarding Form",
        "Added Emergency Contact Person Name, 10-digit Phone Number, and Relationship dropdown (Father, Spouse, Brother, Friend)."
    ],
    [
        "2026-08-05", 8,
        "Third-Party Platform Selection Dropdown",
        "Vivek", "Anurag", "Done", "Onboarding Form",
        "Added a single-select dropdown for existing platform experience with options: Uber, Ola, Rapido, or None."
    ],
    [
        "2026-08-04", 9,
        "Super Admin Role in Approval Chain",
        "Vivek", "Anurag", "Done", "Approvals Dashboard",
        "Included Super Admin role in approver selection dropdowns, forward lists, and pending approval queues."
    ],
    [
        "2026-08-04", 10,
        "Approval Revision Instructions Modal",
        "Vivek", "Anurag", "Done", "Approvals Dashboard",
        "Added full revision modal allowing approvers to send back applications with detailed text instructions for corrections."
    ],
    
    # Vehicle Allocation
    [
        "2026-08-04", 11,
        "Transaction Type Selector",
        "Vivek", "Anurag", "Done", "Allocation Form",
        "Added Transaction Type dropdown with options: New Allocation, Reallocation, Rejoining, and Swap."
    ],
    [
        "2026-08-05", 12,
        "Driver Phone & ID Lookup with Auto-Fill",
        "Vivek", "Anurag", "Done", "Allocation Form",
        "Added Driver Fetch button by phone or ID that automatically fills driver name, city, rental plan, contract type, and car model."
    ],
    [
        "2026-08-04", 13,
        "Vehicle Number Format & Autocomplete Suggestions",
        "Vivek", "Anurag", "Done", "Allocation Form",
        "Added auto-capitalizing vehicle number input with live suggestion dropdown from the active fleet database."
    ],
    [
        "2026-08-04", 14,
        "GPS Active Verification Toggle",
        "Vivek", "Anurag", "Done", "Allocation Form",
        "Added GPS Active dropdown status (Yes/No) for verifying GPS device functionality before vehicle handover."
    ],
    [
        "2026-08-05", 15,
        "OLA Negative Balance Input & Proof Photo Upload",
        "Vivek", "Anurag", "Done", "Allocation & Drop-Off Forms",
        "Added OLA Negative Balance amount field (₹) and OLA Balance Proof Photo upload directly under Driver Information."
    ],
    [
        "2026-08-05", 16,
        "Odometer Reading & Mandatory Photo Upload",
        "Vivek", "Anurag", "Done", "Allocation & Drop-Off Forms",
        "Added Odometer Reading input (KM) and mandatory Odometer Photo capture/upload field."
    ],
    [
        "2026-08-05", 17,
        "Car Condition Photos (Left, Right, Front, Back, Battery)",
        "Vivek", "Anurag", "Done", "Allocation & Drop-Off Forms",
        "Added 5 dedicated photo upload slots for Left Side, Right Side, Front Side, Back Side, and Battery Photo."
    ],
    [
        "2026-08-05", 18,
        "Vehicle Accessories Inspection Checklist",
        "Vivek", "Anurag", "Done", "Allocation Form",
        "Added inspection checklist for Jack, Jack Rod, Spanner, Parking Triangle, Fire Extinguisher, Seat Cover, Floor Carpet, and Music System."
    ],
    [
        "2026-08-04", 19,
        "Swap & Reallocation Old Vehicle Return Details",
        "Vivek", "Anurag", "Done", "Allocation Form",
        "Added Old Vehicle Number, Drop-Off Odometer, Drop-Off Remarks, and Returned Vehicle Inspection Checklist for vehicle swap transactions."
    ],
    [
        "2026-08-04", 20,
        "Strict 10-Digit Mobile & Vehicle Number Validation",
        "Vivek", "Anurag", "Done", "Allocation Form",
        "Enforced strict 10-digit mobile number validation and standard state vehicle registration format checking."
    ],

    # Vehicle Drop-Off
    [
        "2026-08-06", 21,
        "Role Access Control for Vehicle Drop-Off Form",
        "Vivek", "Anurag", "Done", "Form Selector Menu",
        "Granted form access permission for Vehicle Drop-Off Form card to the Onboarding Executive role."
    ],
    [
        "2026-08-06", 22,
        "Drop-Off Reason & Drop-Off Location Dropdowns",
        "Vivek", "Anurag", "Done", "Drop-Off Form",
        "Added Drop-Off Reason dropdown (Voluntary Return, Non-payment/Default, Breakdown, Contract Completion) and Drop-Off Location (Hub, Service Station, Customer Address)."
    ],
    [
        "2026-08-06", 23,
        "Drop-Off Odometer & Car Condition Photos",
        "Vivek", "Anurag", "Done", "Drop-Off Form",
        "Added Odometer reading (KM), Odometer photo, and 5 condition photos (LH, RH, Front, Back, Battery) for vehicle returns."
    ],
    [
        "2026-08-06", 24,
        "Financial Settlement & Deposit Refund Status Fields",
        "Vivek", "Anurag", "Done", "Drop-Off Form",
        "Added Pending Dues (₹), Damage Penalty (₹), and Deposit Refund Status dropdown (Pending Assessment, Approved, Deductions, Forfeited)."
    ],
    [
        "2026-08-06", 25,
        "Save as Draft & Saved Drafts Tab",
        "Vivek", "Anurag", "Done", "Drop-Off Form",
        "Added Save as Draft button and a dedicated Saved Drafts tab to store unfinished drop-off entries for completing later."
    ],
    [
        "2026-08-06", 26,
        "Drop-Off Registry Search, Filters & Export CSV",
        "Vivek", "Anurag", "Done", "Drop-Off Form",
        "Added search bar (driver, phone, vehicle, ID), city/reason/time filters, 10 items per page pagination, and Export CSV button."
    ],
    [
        "2026-08-06", 27,
        "Drop-Off Record Edit & Delete REST Endpoints",
        "Vivek", "Anurag", "Done", "Drop-Off Form & Backend API",
        "Added Edit button to load saved records back into form, Delete button with confirmation, and backend PUT/DELETE API endpoints."
    ],
    [
        "2026-08-06", 28,
        "Indian Standard Time (IST) Timestamp Enforcement",
        "Vivek", "Anurag", "Done", "Backend API",
        "Enforced NOW() AT TIME ZONE 'Asia/Kolkata' across all drop-off creation, updates, and drafts."
    ],

    # Walk-in & Partner Visits
    [
        "2026-08-05", 29,
        "Two-Table Architecture (New Walk-Ins & Existing Partners)",
        "Vivek", "Anurag", "Done", "Walk-In Form & Database",
        "Implemented two-table database architecture (july_new_walkins & july_existing_walkins) with UNION registry listing."
    ],
    [
        "2026-08-05", 30,
        "Returning Partner Visit Logger Mode",
        "Vivek", "Anurag", "Done", "Walk-In Form",
        "Added returning partner visit mode that suppresses candidate onboarding fields and logs only visit reasons and notes."
    ],
    [
        "2026-08-05", 31,
        "Master Onboarding Database Phone Auto-Lookup",
        "Vivek", "Anurag", "Done", "Walk-In Form",
        "Automatically fills driver name and city when typing phone number by searching approved master onboarding records."
    ],
    [
        "2026-08-05", 32,
        "Past Visit History View Modal",
        "Vivek", "Anurag", "Done", "Walk-In Form",
        "Added a visit history modal that opens on phone lookup to view all prior visits logged for a partner."
    ],
    [
        "2026-08-05", 33,
        "Clean ID Formatting (N1, N2, E1, E2)",
        "Vivek", "Anurag", "Done", "Walk-In Registry",
        "Formatted IDs into clean N-series for new candidates and E-series for existing partner visits without hyphens."
    ],
    [
        "2026-08-06", 34,
        "Operational Visit Reasons for Existing Partners",
        "Vivek", "Anurag", "Done", "Walk-In Form",
        "Added Deposit Refund Issue, General Service, Accident, Battery Issue, Tyre Issue, DM Meet, Complaint, Hisaab, and Maintenance options."
    ]
]

# Styling definitions for data cells
font_regular = Font(name="Calibri", size=10, color="0F172A")
font_bold_center = Font(name="Calibri", size=10, bold=True, color="0F172A")
font_status = Font(name="Calibri", size=10, bold=True, color="047857")

fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
fill_odd = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0")
)

start_row = 5
for row_idx, row_data in enumerate(data, start=start_row):
    row_fill = fill_odd if row_idx % 2 == 1 else fill_even
    ws.row_dimensions[row_idx].height = 28
    
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = val
        cell.fill = row_fill
        cell.border = thin_border
        
        if col_idx in [1, 2, 4, 5, 6]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
        if col_idx == 6: # Status
            cell.font = font_status
        elif col_idx in [1, 2]: # Date, Item #
            cell.font = font_bold_center
        else:
            cell.font = font_regular

# Column Widths
col_widths = {
    "A": 14, # Date
    "B": 10, # Item #
    "C": 44, # Feature / Requirement Added
    "D": 16, # Requested By
    "E": 18, # Implemented By
    "F": 12, # Status
    "G": 26, # Location
    "H": 65  # Implementation Details
}

for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save file to Downloads folder
wb.save(output_file)
print(f"[OK] Full 34-item Functional Feature Log saved to Downloads folder: {output_file}")
